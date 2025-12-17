import os
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


# =========================
# LÓGICA DE PROCESSAMENTO
# =========================

def extrair_nome_uc(nome_arquivo: str) -> str:
    """
    Devolve apenas o nome da UC (sem o código e sem o sufixo 'Notas').
    Ex.: '1035754 - Banco de Dados Notas.xlsx' -> 'Banco de Dados'
    """
    base = Path(nome_arquivo).stem  # tira .xlsx
    if base.endswith(" Notas"):
        base = base[:-6]  # remove o " Notas"
    if " - " in base:
        return base.split(" - ", 1)[1]
    return base


def formatar_worksheet(ws):
    """
    Formatação básica:
    - Cabeçalho em negrito, fundo cinza claro, centralizado
    - Bordas em toda a tabela
    - Largura de colunas ajustada
    - Formatação condicional em 'Total do Curso':
        >= 60  -> verde
        40–59  -> amarelo
        < 40   -> vermelho
    """
    max_row = ws.max_row
    max_col = ws.max_column

    # Estilos básicos
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_side = Side(border_style="thin", color="000000")
    border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

    # Cabeçalho
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Bordas e alinhamento linhas de dados
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = border
            if isinstance(cell.value, str):
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Ajuste de largura de colunas
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, max_row + 1):
            cell = ws[f"{col_letter}{row}"]
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = max_len + 2

    # Formatação condicional para 'Total do Curso'
    total_col_idx = None
    for cell in ws[1]:
        if str(cell.value).strip() == "Total do Curso":
            total_col_idx = cell.column
            break

    if total_col_idx and max_row >= 2:
        col_letter = get_column_letter(total_col_idx)
        cell_range = f"{col_letter}2:{col_letter}{max_row}"

        # Verde: >= 60
        fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        regra_verde = CellIsRule(
            operator="greaterThanOrEqual",
            formula=["60"],
            stopIfTrue=False,
            fill=fill_verde,
        )

        # Amarelo: entre 40 e 59
        fill_amarelo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        regra_amarelo = CellIsRule(
            operator="between",
            formula=["40", "59"],
            stopIfTrue=False,
            fill=fill_amarelo,
        )

        # Vermelho: < 40
        fill_vermelho = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        regra_vermelho = CellIsRule(
            operator="lessThan",
            formula=["40"],
            stopIfTrue=False,
            fill=fill_vermelho,
        )

        ws.conditional_formatting.add(cell_range, regra_verde)
        ws.conditional_formatting.add(cell_range, regra_amarelo)
        ws.conditional_formatting.add(cell_range, regra_vermelho)


def processar_arquivos(
    lista_arquivos,
    arquivo_saida: str,
    dividir_por_uc=False,
    manter_nome_original=False,
    log_callback=None,
    progress_callback=None,
):
    if not lista_arquivos:
        raise FileNotFoundError("Nenhum arquivo selecionado.")

    if log_callback:
        log_callback(f"Total de arquivos selecionados: {len(lista_arquivos)}")

    linhas_saida = []
    total_arquivos = len(lista_arquivos)

    for idx, caminho in enumerate(lista_arquivos, start=1):
        arquivo = Path(caminho)
        if log_callback:
            log_callback(f"Processando: {arquivo.name}")

        nome_uc = extrair_nome_uc(arquivo.name)

        # lê a primeira aba
        df = pd.read_excel(arquivo, sheet_name=0)

        col_nome = "Nome"
        col_sobrenome = "Sobrenome"
        col_total_original = "Total do curso (Real)"

        for col in [col_nome, col_sobrenome, col_total_original]:
            if col not in df.columns:
                raise ValueError(
                    f"Coluna obrigatória '{col}' não encontrada no arquivo: {arquivo.name}"
                )

        df["Aluno"] = (
            df[col_nome].astype(str).str.strip()
            + " "
            + df[col_sobrenome].astype(str).str.strip()
        )

        df_final = df[["Aluno", col_total_original]].copy()
        df_final = df_final.rename(columns={col_total_original: "Total do Curso"})

        df_final["UC / Relatório"] = nome_uc

        colunas_ordem = [
            "UC / Relatório",
            "Aluno",
            "Total do Curso",
        ]
        df_final = df_final[colunas_ordem]

        if df_final.empty:
            if log_callback:
                log_callback(f"Nenhuma linha de dados em {arquivo.name}; arquivo ignorado.")
            if progress_callback:
                progress_callback(idx, total_arquivos)
            continue

        # guarda nome do arquivo para possível uso no nome da aba
        df_final.attrs["arquivo_origem"] = arquivo.name
        linhas_saida.append(df_final)

        if progress_callback:
            progress_callback(idx, total_arquivos)

    if not linhas_saida:
        raise ValueError("Nenhuma linha gerada.")

    with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
        if dividir_por_uc:
            usados = {}
            for df_final in linhas_saida:
                nome_uc = df_final["UC / Relatório"].iloc[0]
                if manter_nome_original:
                    base_name = Path(df_final.attrs.get("arquivo_origem", nome_uc)).stem
                else:
                    base_name = nome_uc

                contador = usados.get(base_name, 0)
                if contador == 0:
                    sheet_name = base_name[:31]
                else:
                    sufixo = f"_{contador}"
                    sheet_name = f"{base_name[:31 - len(sufixo)]}{sufixo}"
                usados[base_name] = contador + 1

                df_final.drop(columns=["UC / Relatório"]).to_excel(
                    writer, sheet_name=sheet_name, index=False
                )
                ws = writer.sheets[sheet_name]
                formatar_worksheet(ws)
        else:
            df_saida = pd.concat(linhas_saida, ignore_index=True)
            sheet_name = "Consolidado"
            df_saida.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            formatar_worksheet(ws)

    if log_callback:
        log_callback(f"Arquivo gerado: {arquivo_saida}")


# =========================
# INTERFACE GRÁFICA (UI)
# =========================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.iconbitmap("logo.ico")
        self.title("Consolidador de Notas - SENAI")
        self.geometry("750x450")
        self.resizable(False, False)

        # Tema e estilos
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except:
            pass

        self.configure(bg="#f5f5f5")
        style.configure("TFrame", background="#f5f5f5")
        style.configure("TLabel", background="#f5f5f5", font=("Segoe UI", 9))
        style.configure("Title.TLabel", background="#f5f5f5", font=("Segoe UI", 13, "bold"))
        style.configure("SubTitle.TLabel", background="#f5f5f5", font=("Segoe UI", 9, "italic"))
        style.configure("TButton", padding=6, font=("Segoe UI", 9))
        style.configure("TEntry", padding=3)

        self.arquivos_selecionados = []
        self.resumo_arquivos = tk.StringVar(value="Nenhum arquivo selecionado.")
        self.nome_arquivo_saida = tk.StringVar(value="notas_consolidadas.xlsx")
        self.dividir_por_uc = tk.BooleanVar(value=True)
        self.manter_nome_original = tk.BooleanVar(value=True)
        self.progress_var = tk.DoubleVar(value=0)
        self.progress_total = 0
        self.logs = []

        self._montar_layout()

    def _montar_layout(self):
        padding_geral = {"padx": 10, "pady": 5}

        container = ttk.Frame(self)
        container.pack(fill="both", expand=True, padx=10, pady=10)

        # Header
        header = ttk.Frame(container)
        header.grid(row=0, column=0, columnspan=3, sticky="we")

        lbl_titulo = ttk.Label(header, text="Consolidador de Notas", style="Title.TLabel")
        lbl_titulo.pack(anchor="w")

        lbl_subtitulo = ttk.Label(
            header,
            text="Gera planilha consolidada por UC com formatação visual de desempenho.",
            style="SubTitle.TLabel",
        )
        lbl_subtitulo.pack(anchor="w", pady=(0, 5))

        sep = ttk.Separator(container, orient="horizontal")
        sep.grid(row=1, column=0, columnspan=3, sticky="we", pady=(5, 10))

        # Seleção de arquivos
        lbl_arquivos = ttk.Label(container, text="Arquivos de relatórios (.xlsx / .ods):")
        lbl_arquivos.grid(row=2, column=0, sticky="w", **padding_geral)

        entry_arquivos = ttk.Entry(
            container, textvariable=self.resumo_arquivos, width=60, state="readonly"
        )
        entry_arquivos.grid(row=3, column=0, sticky="we", **padding_geral)

        btn_arquivos = ttk.Button(container, text="Selecionar arquivos", command=self.selecionar_arquivos)
        btn_arquivos.grid(row=3, column=1, sticky="we", **padding_geral)

        # Opção de abas por UC
        chk_dividir = ttk.Checkbutton(
            container,
            text="Gerar uma aba por UC (desmarque para planilha única)",
            variable=self.dividir_por_uc,
        )
        chk_dividir.grid(row=4, column=0, columnspan=2, sticky="w", **padding_geral)

        chk_nome_original = ttk.Checkbutton(
            container,
            text="Usar nome original do arquivo para nome da aba (quando dividir por UC)",
            variable=self.manter_nome_original,
        )
        chk_nome_original.grid(row=5, column=0, columnspan=2, sticky="w", **padding_geral)

        # Nome arquivo saída
        lbl_saida = ttk.Label(container, text="Nome do arquivo de saída (.xlsx):")
        lbl_saida.grid(row=6, column=0, sticky="w", **padding_geral)

        entry_saida = ttk.Entry(container, textvariable=self.nome_arquivo_saida, width=60)
        entry_saida.grid(row=7, column=0, sticky="we", **padding_geral)

        btn_processar = ttk.Button(container, text="Processar relatórios", command=self.on_processar)
        btn_processar.grid(row=7, column=1, sticky="we", **padding_geral)

        # Log
        lbl_log = ttk.Label(container, text="Log de execução:")
        lbl_log.grid(row=8, column=0, sticky="w", **padding_geral)
        btn_exportar_log = ttk.Button(container, text="Exportar log", command=self.exportar_log)
        btn_exportar_log.grid(row=8, column=1, sticky="e", padx=(0, 10), pady=5)

        self.txt_log = tk.Text(
            container, height=10, state="disabled", bg="#ffffff", relief="solid", bd=1
        )
        self.txt_log.grid(row=9, column=0, columnspan=2, sticky="nsew", padx=10, pady=(0, 10))

        scrollbar = ttk.Scrollbar(container, orient="vertical", command=self.txt_log.yview)
        scrollbar.grid(row=9, column=2, sticky="ns", pady=(0, 10))
        self.txt_log["yscrollcommand"] = scrollbar.set

        # Progresso
        frame_progress = ttk.Frame(container)
        frame_progress.grid(row=10, column=0, columnspan=3, sticky="we", padx=10, pady=(0, 5))
        lbl_prog = ttk.Label(frame_progress, text="Progresso:")
        lbl_prog.pack(side="left")
        self.lbl_prog_contador = ttk.Label(frame_progress, text="0/0")
        self.lbl_prog_contador.pack(side="right")
        self.progressbar = ttk.Progressbar(frame_progress, variable=self.progress_var, maximum=100)
        self.progressbar.pack(fill="x", expand=True, padx=(5, 5))

        self.lbl_status = ttk.Label(container, text="Pronto.", anchor="w")
        self.lbl_status.grid(row=11, column=0, columnspan=3, sticky="we", padx=10, pady=(0, 5))

        container.rowconfigure(9, weight=1)
        container.columnconfigure(0, weight=1)

    def selecionar_arquivos(self):
        caminhos = filedialog.askopenfilenames(
            title="Selecione os relatórios",
            filetypes=[
                ("Planilhas Excel/ODS", "*.xlsx *.xls *.ods"),
                ("Excel (.xlsx)", "*.xlsx"),
                ("ODS (.ods)", "*.ods"),
                ("Todos os arquivos", "*.*"),
            ]
        )
        if caminhos:
            self.arquivos_selecionados = list(caminhos)
            qtd = len(self.arquivos_selecionados)
            if qtd == 1:
                resumo = Path(self.arquivos_selecionados[0]).name
            else:
                resumo = f"{qtd} arquivos selecionados."
            self.resumo_arquivos.set(resumo)
            self._set_status("Arquivos selecionados.")

    def log(self, mensagem: str):
        self.logs.append(mensagem)
        self.txt_log.configure(state="normal")
        self.txt_log.insert("end", mensagem + "\n")
        self.txt_log.see("end")
        self.txt_log.configure(state="disabled")
        self.update_idletasks()

    def _set_status(self, msg: str):
        self.lbl_status.config(text=msg)
        self.update_idletasks()

    def _reset_progress(self, total: int):
        self.progress_total = max(total, 0)
        self.progress_var.set(0)
        self.lbl_prog_contador.config(text=f"0/{total}")
        self.update_idletasks()

    def _atualizar_progresso(self, atual: int, total: int):
        total = max(total, 1)
        porcentagem = (atual / total) * 100
        self.progress_var.set(porcentagem)
        self.lbl_prog_contador.config(text=f"{atual}/{total}")
        self.update_idletasks()

    def exportar_log(self):
        if not self.logs:
            messagebox.showinfo("Log", "Nenhum log para exportar.")
            return

        caminho = filedialog.asksaveasfilename(
            title="Salvar log",
            defaultextension=".txt",
            filetypes=[("Arquivo de texto", "*.txt"), ("CSV", "*.csv"), ("Todos os arquivos", "*.*")],
            initialfile="log_processamento.txt",
        )
        if not caminho:
            return

        try:
            Path(caminho).write_text("\n".join(self.logs), encoding="utf-8")
            messagebox.showinfo("Log", f"Log salvo em:\n{caminho}")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível salvar o log:\n{e}")

    def _validar_nome_saida(self, nome_saida: str) -> str:
        nome_saida = nome_saida.strip()
        if not nome_saida:
            raise ValueError("Informe o nome do arquivo de saída.")

        # impede uso de caminhos/pastas no campo do nome
        if Path(nome_saida).name != nome_saida:
            raise ValueError("Não use caminhos ou pastas no nome do arquivo de saída.")

        if not nome_saida.lower().endswith(".xlsx"):
            raise ValueError("Extensão inválida. Use apenas .xlsx.")

        base = nome_saida[:-5]  # remove .xlsx
        caracteres_invalidos = r'\\/:*?"<>|'
        if any(ch in base for ch in caracteres_invalidos):
            raise ValueError("O nome do arquivo não pode conter \\ / : * ? \" < > |")

        if len(base) == 0:
            raise ValueError("Informe um nome válido antes da extensão .xlsx.")

        return nome_saida

    def on_processar(self):
        if not self.arquivos_selecionados:
            messagebox.showwarning("Validação", "Selecione pelo menos um arquivo de relatório.")
            return

        try:
            nome_saida = self._validar_nome_saida(self.nome_arquivo_saida.get())
        except ValueError as e:
            messagebox.showwarning("Validação", str(e))
            return

        pasta_base = Path(self.arquivos_selecionados[0]).parent
        caminho_saida = pasta_base / nome_saida

        try:
            # limpa log anterior e reseta progresso
            self.logs.clear()
            self.txt_log.configure(state="normal")
            self.txt_log.delete("1.0", "end")
            self.txt_log.configure(state="disabled")

            total = len(self.arquivos_selecionados)
            self._reset_progress(total)

            self.log("Iniciando processamento...")
            self._set_status("Processando relatórios...")
            processar_arquivos(
                self.arquivos_selecionados,
                str(caminho_saida),
                dividir_por_uc=self.dividir_por_uc.get(),
                manter_nome_original=self.manter_nome_original.get(),
                log_callback=self.log,
                progress_callback=self._atualizar_progresso,
            )
            self.log("Processamento concluído com sucesso.")
            self._set_status("Processamento concluído.")
            messagebox.showinfo("Sucesso", f"Arquivo gerado:\n{caminho_saida}")
        except Exception as e:
            self.log(f"Erro: {e}")
            self._set_status("Erro no processamento.")
            messagebox.showerror("Erro", f"Ocorreu um erro:\n{e}")


if __name__ == "__main__":
    app = App()
    app.mainloop()
