from __future__ import annotations

from pathlib import Path
from typing import Callable, Iterable, Sequence

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def extrair_nome_uc(nome_arquivo: str) -> str:
    """
    Retorna somente o nome da UC (sem código e sem o sufixo 'Notas').
    Exemplo: '1035754 - Banco de Dados Notas.xlsx' -> 'Banco de Dados'
    """
    base = Path(nome_arquivo).stem
    if base.endswith(" Notas"):
        base = base[:-6]
    if " - " in base:
        return base.split(" - ", 1)[1]
    return base


def extrair_nome_curso(df: pd.DataFrame) -> str | None:
    """
    Extrai o código e nome do curso a partir dos dados do relatório.
    Procura por uma coluna com padrão como '1003121 - Operador de Computador'.
    Retorna no formato: '1003121 - Operador de Computador'
    """
    # Procurar em todas as colunas por um padrão que pareça um código de curso
    for coluna in df.columns:
        if df[coluna].dtype == 'object':
            for valor in df[coluna].dropna():
                valor_str = str(valor).strip()
                # Procurar por padrão como "1003121 - Operador de Computador"
                if " - " in valor_str and any(c.isdigit() for c in valor_str.split(" - ")[0]):
                    primeiro_parte = valor_str.split(" - ")[0].strip()
                    if primeiro_parte.isdigit() and len(primeiro_parte) >= 6:
                        # Formato encontrado: código - nome
                        return valor_str
    return None


def _eh_atividade_valida(nome_atividade: str) -> bool:
    """
    Verifica se a atividade deve ser considerada para o cálculo de notas.
    Descarta SCORM, conteúdo do curso, guias, etc.
    """
    if not nome_atividade or pd.isna(nome_atividade):
        return False
    
    nome = str(nome_atividade).strip().lower()
    
    # Palavras-chave que indicam que NÃO é uma atividade avaliativa
    palavras_excluidas = [
        "scorm",
        "conteúdo do curso",
        "guia de",
        "ambientação",
        "navegação",
        "técnicas de estudos",
    ]
    
    for palavra in palavras_excluidas:
        if palavra in nome:
            return False
    
    return True


def _extrair_notas_por_aluno_e_uc(df: pd.DataFrame) -> dict[str, dict[str, dict]]:
    """
    Extrai e processa as notas por aluno e unidade curricular.
    Retorna um dicionário aninhado:
    {
        'aluno': {
            'uc': {
                'total_notas': soma das notas (apenas com data de correção),
                'quantidade_notas': quantidade de atividades com nota,
                'pendencias': lista de atividades com envio mas sem correção
            }
        }
    }
    """
    resultado = {}
    
    for aluno in df['aluno'].unique():
        if pd.isna(aluno):
            continue
        
        df_aluno = df[df['aluno'] == aluno]
        resultado[aluno] = {}
        
        # Agrupar por unidade curricular
        for uc in df_aluno['unidade curricular'].unique():
            if pd.isna(uc):
                continue
            
            df_uc = df_aluno[df_aluno['unidade curricular'] == uc]
            
            total_notas = 0
            quantidade_notas = 0
            pendencias = []
            
            for _, row in df_uc.iterrows():
                atividade = row.get('atividade')
                nota = row.get('nota final')
                data_envio = row.get('data do envio')
                data_correcao = row.get('data de correção')
                
                # Verifica se é uma atividade válida
                if not _eh_atividade_valida(atividade):
                    continue
                
                # Converter para strings para verificação
                tem_envio = not pd.isna(data_envio) and str(data_envio).strip() != ''
                tem_correcao = not pd.isna(data_correcao) and str(data_correcao).strip() != ''
                
                # Se tem envio mas não tem correção, adiciona às pendências
                if tem_envio and not tem_correcao:
                    pendencias.append({
                        'atividade': atividade,
                        'data_envio': data_envio,
                    })
                
                # Apenas conta notas com data de correção e nota >= 2
                if tem_correcao and not pd.isna(nota) and nota >= 2:
                    total_notas += nota
                    quantidade_notas += 1
            
            resultado[aluno][uc] = {
                'total_notas': total_notas,
                'quantidade_notas': quantidade_notas,
                'pendencias': pendencias,
            }
    
    return resultado


def formatar_worksheet(ws) -> None:
    """
    Formatação básica:
    - Cabeçalho em negrito, fundo azul escuro, texto branco
    - Bordas em toda a tabela
    - Largura de colunas ajustada
    - Formatação condicional em todas as colunas numéricas
    """
    max_row = ws.max_row
    max_col = ws.max_column

    # Estilos básicos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="366092")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side = Side(border_style="thin", color="000000")
    border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

    # Cabeçalho
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Bordas e alinhamento das linhas de dados
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = border
            if isinstance(cell.value, (str, type(None))):
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Ajuste de largura de colunas
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, max_row + 1):
            cell = ws[f"{col_letter}{row}"]
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Formatação condicional para todas as colunas numéricas (exceto "Aluno")
    regra_verde = CellIsRule(
        operator="greaterThanOrEqual",
        formula=["60"],
        stopIfTrue=False,
        fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    )
    regra_amarela = CellIsRule(
        operator="between",
        formula=["40", "59"],
        stopIfTrue=False,
        fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    )
    regra_vermelha = CellIsRule(
        operator="lessThan",
        formula=["40"],
        stopIfTrue=False,
        fill=PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),
    )

    # Aplicar formatação condicional a todas as colunas de UC (coluna 2 em diante)
    for col_idx in range(2, max_col + 1):
        col_letter = get_column_letter(col_idx)
        # Range de dados (sem o cabeçalho)
        cell_range = f"{col_letter}2:{col_letter}{max_row}"
        
        ws.conditional_formatting.add(cell_range, regra_verde)
        ws.conditional_formatting.add(cell_range, regra_amarela)
        ws.conditional_formatting.add(cell_range, regra_vermelha)


def processar_arquivos(
    lista_arquivos: Sequence[str],
    arquivo_saida: str,
    *,
    dividir_por_uc: bool = False,
    log_callback: Callable[[str], None] | None = None,
    progress_callback: Callable[[int, int], None] | None = None,
) -> None:
    """
    Consolida os relatórios em um único arquivo Excel.
    
    Funcionalidades:
    - Filtra atividades válidas (descarта SCORM, conteúdo do curso, etc.)
    - Considera apenas notas >= 2 e com data de correção
    - Gera lista de pendências (envios sem correção)
    - Cria abas separadas para dados e pendências

    Args:
        lista_arquivos: caminhos dos relatórios .xlsx/.ods.
        arquivo_saida: caminho completo do arquivo consolidado.
        dividir_por_uc: cria uma aba por arquivo/turma quando True.
        log_callback: função para registrar mensagens no UI.
        progress_callback: função para atualizar progresso (atual, total).
    """
    if not lista_arquivos:
        raise FileNotFoundError("Nenhum arquivo selecionado.")

    if log_callback:
        log_callback(f"Total de arquivos selecionados: {len(lista_arquivos)}")

    linhas_saida: list[pd.DataFrame] = []
    pendencias_por_uc: dict[str, list[dict]] = {}
    total_arquivos = len(lista_arquivos)

    for idx, caminho in enumerate(lista_arquivos, start=1):
        arquivo = Path(caminho)
        if log_callback:
            log_callback(f"Processando: {arquivo.name}")

        nome_uc = extrair_nome_uc(arquivo.name)
        df = pd.read_excel(arquivo, sheet_name=0)

        # Validar colunas necessárias
        colunas_obrigatorias = ["aluno", "atividade", "nota final", "data do envio", "data de correção", "unidade curricular"]
        for col in colunas_obrigatorias:
            if col not in df.columns:
                raise ValueError(f"Coluna obrigatória '{col}' não encontrada em: {arquivo.name}")

        # Extrair notas por aluno e UC
        dados_alunos = _extrair_notas_por_aluno_e_uc(df)

        # Preparar dataframe de saída (formato pivotado: Aluno x UC)
        linhas_dados = []
        todas_ucs = set()
        
        for aluno, ucs_dados in dados_alunos.items():
            linha = {'Aluno': aluno}
            
            # Coletar todas as UCs
            for uc in ucs_dados.keys():
                todas_ucs.add(uc)
            
            # Adicionar totais de cada UC
            for uc, dados in ucs_dados.items():
                linha[uc] = dados['total_notas']
                
                # Acumular pendências por UC
                if dados['pendencias']:
                    if uc not in pendencias_por_uc:
                        pendencias_por_uc[uc] = []
                    
                    for pendencia in dados['pendencias']:
                        pendencias_por_uc[uc].append({
                            'Aluno': aluno,
                            'Atividade': pendencia['atividade'],
                            'Data de Envio': pendencia['data_envio'],
                        })
            
            linhas_dados.append(linha)

        if linhas_dados:
            # Converter para DataFrame e ordenar colunas (Aluno primeiro, depois UCs)
            df_final = pd.DataFrame(linhas_dados)
            
            # Ordenar UCs
            todas_ucs_ordenadas = sorted(todas_ucs)
            colunas_ordem = ['Aluno'] + todas_ucs_ordenadas
            df_final = df_final[colunas_ordem]
            
            # Adicionar linha de médias
            linha_medias = {'Aluno': 'MÉDIA'}
            for uc in todas_ucs_ordenadas:
                linha_medias[uc] = df_final[uc].mean()
            
            df_final = pd.concat([df_final, pd.DataFrame([linha_medias])], ignore_index=True)
            
            # Salvar atributos para uso posterior
            df_final.attrs["arquivo_origem"] = arquivo.name
            nome_curso = extrair_nome_curso(df)
            if nome_curso:
                df_final.attrs["nome_curso"] = nome_curso
            
            linhas_saida.append(df_final)

            if log_callback:
                log_callback(f"  {len(linhas_dados)} aluno(s) e {len(todas_ucs)} UC(s) processado(s)")
        else:
            if log_callback:
                log_callback(f"  Nenhuma linha de dados válida em {arquivo.name}")

        if progress_callback:
            progress_callback(idx, total_arquivos)

    if not linhas_saida:
        raise ValueError("Nenhuma linha gerada após processamento.")

    # Função auxiliar para gerar nome de aba válido (máx 31 caracteres Excel)
    def gerar_nome_aba(nome: str, max_len: int = 31) -> str:
        """Gera nome de aba truncado se necessário mantendo legibilidade."""
        if len(nome) <= max_len:
            return nome
        
        # Tentar truncar no último espaço antes do limite
        nome_truncado = nome[:max_len]
        ultimo_espaco = nome_truncado.rfind(' ')
        
        if ultimo_espaco > 10:  # Garante mínimo de caracteres
            return nome[:ultimo_espaco].rstrip() + "…"
        else:
            # Se não tiver espaço bom, truncar simples com reticências
            return nome[:max_len-1] + "…"

    # Salvar no Excel
    with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
        if dividir_por_uc:
            # Dividir dados por arquivo/turma em abas separadas
            # Cada arquivo (turma) terá sua própria aba com seus dados
            for idx, df_final in enumerate(linhas_saida, start=1):
                # Preferir nome do curso, senão usar nome do arquivo
                nome_aba = df_final.attrs.get("nome_curso")
                
                if not nome_aba:
                    # Fallback: usar nome do arquivo (sem extensão)
                    arquivo_origem = df_final.attrs.get("arquivo_origem", f"Turma {idx}")
                    nome_aba = Path(arquivo_origem).stem
                
                # Truncar se necessário
                nome_aba = gerar_nome_aba(nome_aba)
                
                # Evitar nomes duplicados adicionando índice se necessário
                nome_aba_final = nome_aba
                contador = 1
                while nome_aba_final in writer.sheets:
                    nome_aba_final = gerar_nome_aba(f"{nome_turma} ({contador})")
                    contador += 1
                
                df_final.to_excel(writer, sheet_name=nome_aba_final, index=False)
                ws = writer.sheets[nome_aba_final]
                formatar_worksheet(ws)
        else:
            # Formato consolidado: uma única aba com todos os dados
            df_saida = pd.concat(linhas_saida, ignore_index=True)
            sheet_name = "Notas"
            df_saida.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            formatar_worksheet(ws)

        # Gerar aba de pendências
        if pendencias_por_uc:
            linhas_pendencias = []
            for uc, pendencias in pendencias_por_uc.items():
                for pend in pendencias:
                    pend_com_uc = pend.copy()
                    pend_com_uc['UC'] = uc
                    linhas_pendencias.append(pend_com_uc)
            
            if linhas_pendencias:
                df_pendencias = pd.DataFrame(linhas_pendencias)
                colunas_ordem = ["UC", "Aluno", "Atividade", "Data de Envio"]
                df_pendencias = df_pendencias[colunas_ordem]
                df_pendencias.to_excel(writer, sheet_name="Pendências", index=False)
                ws = writer.sheets["Pendências"]
                formatar_worksheet(ws)
                
                if log_callback:
                    log_callback(f"\nTotal de atividades pendentes: {len(linhas_pendencias)}")

    if log_callback:
        log_callback(f"\nArquivo gerado: {arquivo_saida}")

